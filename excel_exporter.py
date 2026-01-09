# excel_exporter.py
import re
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.filters import (
    AutoFilter, FilterColumn, CustomFilter, Filters
)
from typing import List, Dict

def export_to_excel(posts: List[Dict], output_path: str):
    if not posts:
        print("No posts to export")
        return
    
    # Define column order - ADDED "Date" column
    columns = [
        "Slide", "Post #", "Type", "Title", "Date",
        "Reach", "Views", "Engagement", 
        "Likes", "Shares", "Comments", "Saved",
        "Link"
    ]
    
    # Create DataFrame with separated title and date
    rows = []
    for post in posts:
        original_title = post.get("title", "")
        
        # Extract date from title
        post_date = ""
        clean_title = original_title
        
        # Look for date pattern [dd/mm]
        date_match = re.search(r'\[(\d{1,2}/\d{1,2})\]', original_title)
        if date_match:
            post_date = date_match.group(1)
            # Remove date from title
            clean_title = original_title.replace(f"[{post_date}]", "").strip()
            # Clean up extra spaces
            clean_title = re.sub(r'\s+', ' ', clean_title)
        
        rows.append({
            "Slide": post.get("slide_number"),
            "Post #": post.get("post_index"),
            "Type": post.get("type", "").capitalize(),
            "Title": clean_title,
            "Date": post_date,
            "Reach": post.get("reach"),
            "Views": post.get("views"),
            "Engagement": post.get("engagement"),
            "Likes": post.get("likes"),
            "Shares": post.get("shares"),
            "Comments": post.get("comments"),
            "Saved": post.get("saved"),
            "Link": post.get("link", ""),
        })
    
    df = pd.DataFrame(rows, columns=columns)
    
    # Create Excel writer
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Social Media Posts')
        
        # Get workbook and worksheet
        workbook = writer.book
        worksheet = writer.sheets['Social Media Posts']
        
        # ========== DEFINE STYLES ==========
        # Header styles
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")  # Dark blue
        center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        # Data styles
        thin_border = Border(
            left=Side(style='thin', color='D3D3D3'),
            right=Side(style='thin', color='D3D3D3'),
            top=Side(style='thin', color='D3D3D3'),
            bottom=Side(style='thin', color='D3D3D3')
        )
        
        number_alignment = Alignment(horizontal="center", vertical="center")
        text_alignment = Alignment(vertical="center", wrap_text=True)
        link_alignment = Alignment(vertical="center", wrap_text=True)
        
        # Type-specific colors
        type_colors = {
            "Video": PatternFill(start_color="E8F4FD", end_color="E8F4FD", fill_type="solid"),  # Light blue
            "Post": PatternFill(start_color="F0F8FF", end_color="F0F8FF", fill_type="solid"),   # Very light blue
            "Ad": PatternFill(start_color="FFF0F5", end_color="FFF0F5", fill_type="solid"),     # Lavender blush
            "Community": PatternFill(start_color="F0FFF0", end_color="F0FFF0", fill_type="solid")  # Honeydew
        }
        
        # ========== APPLY HEADER FORMATTING ==========
        for col in range(1, len(columns) + 1):
            cell = worksheet.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
            cell.border = thin_border
        
        # ========== APPLY DATA FORMATTING ==========
        for row in range(2, len(df) + 2):
            for col_idx, col_name in enumerate(columns, start=1):
                cell = worksheet.cell(row=row, column=col_idx)
                cell.border = thin_border
                
                # Format based on column type
                if col_name in ["Reach", "Views", "Engagement", "Likes", "Shares", "Comments", "Saved"]:
                    cell.alignment = number_alignment
                    if cell.value is not None:
                        cell.number_format = '#,##0'
                        # Add conditional formatting for high values - with type checking
                        try:
                            # Convert to number if it's a string
                            if isinstance(cell.value, str):
                                # Remove commas and convert
                                numeric_value = float(cell.value.replace(',', ''))
                            else:
                                numeric_value = float(cell.value)
                            
                            if numeric_value > 10000:
                                cell.font = Font(bold=True, color="006400")  # Dark green for high values
                        except (ValueError, TypeError):
                            # If conversion fails, skip the formatting
                            pass
                elif col_name == "Type":
                    cell.alignment = center_alignment
                    # Apply color coding based on post type
                    if cell.value in type_colors:
                        cell.fill = type_colors[cell.value]
                        cell.font = Font(bold=True)
                elif col_name == "Title":
                    cell.alignment = text_alignment
                elif col_name == "Link":
                    cell.alignment = link_alignment
                    # Make links clickable
                    if cell.value and isinstance(cell.value, str) and cell.value.startswith("http"):
                        cell.hyperlink = cell.value
                        cell.font = Font(color="0563C1", underline="single")
                elif col_name in ["Slide", "Post #"]:
                    cell.alignment = center_alignment
        
        # ========== ADD CONDITIONAL FORMATTING ==========
        # Color scale for engagement metrics
        engagement_col_idx = columns.index("Engagement") + 1
        engagement_col_letter = get_column_letter(engagement_col_idx)
        engagement_range = f"{engagement_col_letter}2:{engagement_col_letter}{len(df) + 1}"
        
        # 3-color scale for engagement
        color_scale_rule = ColorScaleRule(
            start_type='min', start_color='FFF6F6',
            mid_type='percentile', mid_color='FFEB84',
            end_type='max', end_color='63BE7B'
        )
        worksheet.conditional_formatting.add(engagement_range, color_scale_rule)
        
        # Highlight videos (if "Video" in Type column)
        type_col_idx = columns.index("Type") + 1
        type_col_letter = get_column_letter(type_col_idx)
        type_range = f"{type_col_letter}2:{type_col_letter}{len(df) + 1}"
        
        video_rule = CellIsRule(
            operator='equal', 
            formula=['"Video"'], 
            fill=PatternFill(start_color='E1F5FE', end_color='E1F5FE', fill_type='solid')
        )
        worksheet.conditional_formatting.add(type_range, video_rule)
        
        # Highlight posts with no engagement
        low_engagement_rule = CellIsRule(
            operator='lessThanOrEqual',
            formula=['100'],
            font=Font(color='FF6B6B', italic=True)
        )
        worksheet.conditional_formatting.add(engagement_range, low_engagement_rule)
        
        # ========== ADD FILTERS ==========
        # Enable auto-filter for all columns
        worksheet.auto_filter.ref = f"A1:{get_column_letter(len(columns))}1"
        
        # ========== ADD SORTING BUTTONS (via Freeze Panes with header) ==========
        worksheet.freeze_panes = 'A2'  # Freeze header row
        
        # ========== ADD SUMMARY STATISTICS ==========
        summary_row = len(df) + 4
        
        # Summary header
        worksheet.cell(row=summary_row, column=1, value="📊 SUMMARY STATISTICS").font = Font(bold=True, size=12, color="2C3E50")
        summary_row += 1
        
        # Calculate statistics
        total_posts = len(df)
        video_count = len(df[df['Type'] == 'Video'])
        avg_reach = df['Reach'].mean() if not df['Reach'].isna().all() else 0
        avg_engagement = df['Engagement'].mean() if not df['Engagement'].isna().all() else 0
        total_engagement = df['Engagement'].sum() if not df['Engagement'].isna().all() else 0
        
        # Add summary data
        stats = [
            ("Total Posts", total_posts),
            ("Video Posts", video_count),
            ("Average Reach", f"{avg_reach:,.0f}"),
            ("Average Engagement", f"{avg_engagement:,.0f}"),
            ("Total Engagement", f"{total_engagement:,.0f}"),
            ("Post Types", df['Type'].value_counts().to_dict())
        ]
        
        for i, (label, value) in enumerate(stats):
            if label == "Post Types":
                worksheet.cell(row=summary_row + i, column=1, value=label).font = Font(bold=True)
                type_dict = value
                for j, (type_name, count) in enumerate(type_dict.items()):
                    worksheet.cell(row=summary_row + i, column=2 + j, value=f"{type_name}: {count}").font = Font(bold=True)
            else:
                worksheet.cell(row=summary_row + i, column=1, value=label).font = Font(bold=True)
                worksheet.cell(row=summary_row + i, column=2, value=value)
        
        # ========== ADD CHARTS SHEET ==========
        if total_posts > 0:
            try:
                from openpyxl.chart import BarChart, Reference, Series
                
                # Create a new sheet for charts
                charts_sheet = workbook.create_sheet(title="Charts")
                
                # Prepare data for chart
                chart_data = []
                
                # Chart 1: Posts by Type
                type_counts = df['Type'].value_counts()
                if not type_counts.empty:
                    chart1_data = pd.DataFrame({
                        'Type': type_counts.index,
                        'Count': type_counts.values
                    })
                    
                    # Write chart data
                    charts_sheet['A1'] = "Posts by Type"
                    charts_sheet['A1'].font = Font(bold=True, size=14)
                    
                    for i, (type_name, count) in enumerate(zip(chart1_data['Type'], chart1_data['Count']), start=3):
                        charts_sheet[f'A{i}'] = type_name
                        charts_sheet[f'B{i}'] = count
                    
                    # Create chart
                    chart1 = BarChart()
                    chart1.type = "col"
                    chart1.style = 10
                    chart1.title = "Posts by Type"
                    chart1.y_axis.title = 'Number of Posts'
                    chart1.x_axis.title = 'Post Type'
                    
                    data = Reference(charts_sheet, min_col=2, min_row=2, max_row=2 + len(type_counts))
                    cats = Reference(charts_sheet, min_col=1, min_row=3, max_row=3 + len(type_counts))
                    chart1.add_data(data, titles_from_data=True)
                    chart1.set_categories(cats)
                    
                    charts_sheet.add_chart(chart1, "D2")
                
                # Chart 2: Top 10 Posts by Engagement
                top_posts = df.nlargest(10, 'Engagement')[['Title', 'Engagement']]
                if not top_posts.empty:
                    charts_sheet['A15'] = "Top 10 Posts by Engagement"
                    charts_sheet['A15'].font = Font(bold=True, size=14)
                    
                    for i, (_, row) in enumerate(top_posts.iterrows(), start=17):
                        title = row['Title'][:30] + "..." if len(row['Title']) > 30 else row['Title']
                        charts_sheet[f'A{i}'] = title
                        charts_sheet[f'B{i}'] = row['Engagement']
                    
                    chart2 = BarChart()
                    chart2.type = "col"
                    chart2.style = 11
                    chart2.title = "Top 10 Posts by Engagement"
                    chart2.y_axis.title = 'Engagement'
                    
                    data = Reference(charts_sheet, min_col=2, min_row=16, max_row=16 + len(top_posts))
                    cats = Reference(charts_sheet, min_col=1, min_row=17, max_row=17 + len(top_posts))
                    chart2.add_data(data, titles_from_data=True)
                    chart2.set_categories(cats)
                    
                    charts_sheet.add_chart(chart2, "D15")
                    
            except ImportError:
                pass  # Charts are optional
        
        # ========== AUTO-ADJUST COLUMN WIDTHS ==========
        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            
            # Set width with different minimums based on column type
            col_name = columns[column[0].column - 1]
            if col_name == "Title":
                adjusted_width = min(max(max_length + 2, 25), 60)
            elif col_name == "Link":
                adjusted_width = min(max(max_length + 2, 20), 50)
            elif col_name in ["Reach", "Views", "Engagement"]:
                adjusted_width = 15
            elif col_name in ["Likes", "Shares", "Comments", "Saved"]:
                adjusted_width = 12
            else:
                adjusted_width = min(max_length + 2, 20)
            
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # ========== ADD DATA VALIDATION FOR SORTING ==========
        # Create a separate sheet for sorting instructions
        instructions_sheet = workbook.create_sheet(title="Instructions")
        
        instructions = [
            ("📋 INSTRUCTIONS", ""),
            ("", ""),
            ("📊 Filters:", "Click the dropdown arrows in the header row to filter data"),
            ("", ""),
            ("🔍 Sort by Column:", "Click any column header to sort A-Z or Z-A"),
            ("", ""),
            ("🎨 Color Coding:", ""),
            ("  • Videos:", "Light blue background"),
            ("  • High Engagement (>10K):", "Bold green text"),
            ("  • Low Engagement (<100):", "Red italic text"),
            ("", ""),
            ("📈 Charts:", "See 'Charts' sheet for visual summaries"),
            ("", ""),
            ("💡 Tips:", ""),
            ("  • Use 'Type' filter to view only Videos/Posts/Ads", ""),
            ("  • Sort by 'Engagement' to find top performers", ""),
            ("  • Click links to open social media posts", ""),
        ]
        
        for i, (left, right) in enumerate(instructions, start=1):
            instructions_sheet.cell(row=i, column=1, value=left)
            instructions_sheet.cell(row=i, column=3, value=right)
            
            if left.startswith("📋"):
                instructions_sheet.cell(row=i, column=1).font = Font(bold=True, size=14, color="2C3E50")
            elif left.startswith("📊") or left.startswith("🔍") or left.startswith("🎨") or left.startswith("📈") or left.startswith("💡"):
                instructions_sheet.cell(row=i, column=1).font = Font(bold=True, color="2C3E50")
        
        # Auto-adjust instruction column widths
        instructions_sheet.column_dimensions['A'].width = 30
        instructions_sheet.column_dimensions['C'].width = 50
        
        # Move Instructions sheet to front
        workbook._sheets.insert(0, workbook._sheets.pop(workbook._sheets.index(instructions_sheet)))
        
    print(f"✅ Saved {len(posts)} items to {output_path}")
    print(f"   Features: Filters, Sorting, Color Coding, Summary Stats, Charts")