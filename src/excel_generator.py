"""
Enhanced Excel Generator with advanced formatting and analytics
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, PieChart, Reference, Series
from openpyxl.chart.label import DataLabelList
from openpyxl.drawing.image import Image
from typing import Dict, List, Any, Optional
import os
from pathlib import Path
from datetime import datetime
import logging

class EnhancedExcelGenerator:
    """Generate comprehensive Excel reports with advanced analytics"""
    
    def __init__(self):
        self.initialize_styles()
        self.logger = logging.getLogger(__name__)
        
    def initialize_styles(self):
        """Initialize Excel styles"""
        # Colors
        self.primary_color = "366092"  # Dark Blue
        self.secondary_color = "4F81BD"  # Light Blue
        self.success_color = "00B050"  # Green
        self.warning_color = "FFC000"  # Yellow
        self.danger_color = "FF0000"  # Red
        self.light_gray = "F2F2F2"
        self.dark_gray = "D9D9D9"
        
        # Header styles
        self.header_fill = PatternFill(
            start_color=self.primary_color,
            end_color=self.primary_color,
            fill_type="solid"
        )
        self.header_font = Font(color="FFFFFF", bold=True, size=11, name="Calibri")
        
        # Title styles
        self.title_font = Font(bold=True, size=16, color=self.primary_color, name="Calibri")
        self.subtitle_font = Font(bold=True, size=12, color=self.secondary_color, name="Calibri")
        
        # Data styles
        self.data_font = Font(size=10, name="Calibri")
        self.bold_font = Font(bold=True, size=10, name="Calibri")
        self.number_font = Font(size=10, name="Calibri")
        self.currency_font = Font(size=10, name="Calibri")
        
        # Borders
        self.thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        self.thick_border = Border(
            left=Side(style='medium'),
            right=Side(style='medium'),
            top=Side(style='medium'),
            bottom=Side(style='medium')
        )
        
        # Alignments
        self.center_alignment = Alignment(horizontal='center', vertical='center')
        self.left_alignment = Alignment(horizontal='left', vertical='center')
        self.right_alignment = Alignment(horizontal='right', vertical='center')
        
        # Fills
        self.alt_row_fill = PatternFill(
            start_color=self.light_gray,
            end_color=self.light_gray,
            fill_type="solid"
        )
        
        self.highlight_fill = PatternFill(
            start_color="FFF2CC",
            end_color="FFF2CC",
            fill_type="solid"
        )
    
    def generate_dashboard_report(self, all_data: Dict[str, List], 
                                 output_path: str,
                                 include_charts: bool = True,
                                 include_summary: bool = True) -> str:
        """
        Generate comprehensive Excel report
        
        Args:
            all_data: Dictionary of data by dashboard type
            output_path: Output Excel file path
            include_charts: Whether to include charts
            include_summary: Whether to include summary dashboard
            
        Returns:
            Path to generated Excel file
        """
        start_time = datetime.now()
        
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            workbook = writer.book
            
            # Create cover sheet
            self._create_cover_sheet(workbook, all_data)
            
            # Create summary dashboard if requested
            if include_summary:
                self._create_summary_dashboard(workbook, all_data)
            
            # Create sheets for each dashboard type
            for dashboard_type, data_list in all_data.items():
                if data_list:
                    df = pd.DataFrame(data_list)
                    sheet_name = self._format_sheet_name(dashboard_type)
                    
                    # Write data
                    df.to_excel(writer, sheet_name=sheet_name, index=False)
                    
                    # Get worksheet
                    worksheet = writer.sheets[sheet_name]
                    
                    # Apply formatting
                    self._format_dashboard_sheet(worksheet, df, dashboard_type)
                    
                    # Add summary statistics
                    self._add_dashboard_summary(worksheet, df, dashboard_type)
                    
                    # Add charts if requested
                    if include_charts:
                        self._add_charts_to_sheet(worksheet, df, dashboard_type)
            
            # Create data quality report
            self._create_data_quality_sheet(workbook, all_data)
            
            # Create executive summary
            self._create_executive_summary(workbook, all_data)
            
            # Adjust column widths
            for sheet in workbook.worksheets:
                self._auto_adjust_columns(sheet)
        
        processing_time = (datetime.now() - start_time).total_seconds()
        self.logger.info(f"Excel report generated in {processing_time:.2f} seconds: {output_path}")
        
        return output_path
    
    def _create_cover_sheet(self, workbook, all_data: Dict[str, List]):
        """Create professional cover sheet"""
        ws = workbook.create_sheet(title="Cover", index=0)
        
        # Add logo/title area
        ws.merge_cells('A1:H3')
        ws['A1'] = "MARKETING PERFORMANCE DASHBOARD"
        ws['A1'].font = self.title_font
        ws['A1'].alignment = self.center_alignment
        
        # Add subtitle
        ws.merge_cells('A4:H5')
        ws['A4'] = "Comprehensive Analytics Report"
        ws['A4'].font = self.subtitle_font
        ws['A4'].alignment = self.center_alignment
        
        # Add report info
        ws['A7'] = "Report Generated:"
        ws['B7'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        ws['A7'].font = self.bold_font
        ws['B7'].font = self.data_font
        
        ws['A8'] = "Data Sources:"
        ws['B8'] = "PPT Files Analysis"
        ws['A8'].font = self.bold_font
        ws['B8'].font = self.data_font
        
        # Add dashboard summary table
        ws['A10'] = "Dashboard Summary"
        ws['A10'].font = Font(bold=True, size=14, color=self.primary_color)
        ws.merge_cells('A10:H10')
        
        # Table headers
        headers = ["Dashboard", "Records", "Files", "Date Range", "Status", "Quality"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=12, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.center_alignment
            cell.border = self.thin_border
        
        # Table data
        row = 13
        total_records = 0
        
        for dashboard_type, data_list in all_data.items():
            if data_list:
                df = pd.DataFrame(data_list)
                records = len(data_list)
                total_records += records
                
                # Calculate metrics
                files = df['source_file'].nunique() if 'source_file' in df.columns else 1
                date_range = self._get_date_range_from_df(df)
                
                # Data quality assessment
                quality_score = self._calculate_data_quality(df, dashboard_type)
                quality_status = self._get_quality_status(quality_score)
                
                values = [
                    self._format_dashboard_name(dashboard_type),
                    records,
                    files,
                    date_range,
                    "✓ Complete",
                    quality_status
                ]
                
                for col, value in enumerate(values, 1):
                    cell = ws.cell(row=row, column=col, value=value)
                    cell.font = self.data_font
                    cell.border = self.thin_border
                    cell.alignment = self.center_alignment
                    
                    # Alternate row coloring
                    if row % 2 == 0:
                        cell.fill = self.alt_row_fill
                    
                    # Color code quality status
                    if col == 6:  # Quality column
                        if quality_status == "Excellent":
                            cell.font = Font(color=self.success_color, bold=True)
                        elif quality_status == "Good":
                            cell.font = Font(color=self.success_color)
                        elif quality_status == "Fair":
                            cell.font = Font(color=self.warning_color)
                        else:
                            cell.font = Font(color=self.danger_color)
                
                row += 1
        
        # Add total row
        ws.cell(row=row, column=1, value="TOTAL").font = self.bold_font
        ws.cell(row=row, column=2, value=total_records).font = self.bold_font
        ws.cell(row=row, column=2).fill = self.highlight_fill
        
        # Add instructions
        instruction_row = row + 2
        ws.merge_cells(f'A{instruction_row}:H{instruction_row+1}')
        instruction_cell = ws.cell(row=instruction_row, column=1)
        instruction_cell.value = "Navigate to individual dashboard sheets for detailed analysis and charts"
        instruction_cell.font = Font(italic=True, color=self.secondary_color, size=11)
        instruction_cell.alignment = self.center_alignment
        
        # Add sheet navigation guide
        guide_row = instruction_row + 3
        ws.merge_cells(f'A{guide_row}:H{guide_row}')
        guide_cell = ws.cell(row=guide_row, column=1)
        guide_cell.value = "Sheet Navigation Guide"
        guide_cell.font = Font(bold=True, size=12, color=self.primary_color)
        guide_cell.alignment = self.center_alignment
        
        guide_items = [
            ("Summary", "High-level overview and key metrics"),
            ("Social_Media", "Platform performance and engagement analysis"),
            ("Performance_Marketing", "Ad campaign ROI and efficiency"),
            ("KOL_Engagement", "Influencer performance and tier analysis"),
            ("Data_Quality", "Data validation and quality assessment"),
            ("Executive_Summary", "Strategic insights and recommendations")
        ]
        
        for i, (sheet_name, description) in enumerate(guide_items):
            ws.cell(row=guide_row + 2 + i, column=1, value=f"• {sheet_name}:").font = self.bold_font
            ws.cell(row=guide_row + 2 + i, column=2, value=description).font = self.data_font
    
    def _create_summary_dashboard(self, workbook, all_data: Dict[str, List]):
        """Create summary dashboard with charts"""
        ws = workbook.create_sheet(title="Summary", index=1)
        
        # Title
        ws.merge_cells('A1:H1')
        ws['A1'] = "EXECUTIVE DASHBOARD - KEY METRICS"
        ws['A1'].font = Font(bold=True, size=18, color=self.primary_color)
        ws['A1'].alignment = self.center_alignment
        
        # Subtitle
        ws.merge_cells('A2:H2')
        ws['A2'] = "Automated Marketing Performance Analysis"
        ws['A2'].font = Font(size=12, color=self.secondary_color)
        ws['A2'].alignment = self.center_alignment
        
        # Create metrics summary
        metrics_data = []
        total_engagement = 0
        
        for dashboard_type, data_list in all_data.items():
            if data_list:
                df = pd.DataFrame(data_list)
                records = len(data_list)
                engagement = self._calculate_total_engagement(df)
                total_engagement += engagement
                
                metrics_data.append({
                    'dashboard': self._format_dashboard_name(dashboard_type),
                    'records': records,
                    'engagement': engagement,
                    'avg_engagement': engagement / records if records > 0 else 0
                })
        
        # Write metrics summary table
        ws['A4'] = "Dashboard Performance Summary"
        ws['A4'].font = Font(bold=True, size=14, color=self.primary_color)
        ws.merge_cells('A4:E4')
        
        headers = ["Dashboard", "Records", "Total Engagement", "Avg/Record", "Share %"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=6, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.center_alignment
        
        row = 7
        for data in metrics_data:
            share_percentage = (data['engagement'] / total_engagement * 100) if total_engagement > 0 else 0
            
            values = [
                data['dashboard'],
                data['records'],
                data['engagement'],
                data['avg_engagement'],
                f"{share_percentage:.1f}%"
            ]
            
            for col, value in enumerate(values, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.font = self.data_font
                cell.border = self.thin_border
                cell.alignment = self.center_alignment
            
            row += 1
        
        # Add charts
        chart_row = row + 2
        
        # Bar chart for records by dashboard
        chart1 = BarChart()
        chart1.type = "col"
        chart1.style = 10
        chart1.title = "Records by Dashboard"
        chart1.y_axis.title = "Number of Records"
        chart1.x_axis.title = "Dashboard"
        
        data = Reference(ws, min_col=2, min_row=6, max_row=row-1)
        categories = Reference(ws, min_col=1, min_row=7, max_row=row-1)
        
        chart1.add_data(data, titles_from_data=False)
        chart1.set_categories(categories)
        chart1.shape = 4
        
        ws.add_chart(chart1, f"A{chart_row}")
        
        # Pie chart for engagement distribution
        chart2 = PieChart()
        chart2.title = "Engagement Distribution"
        
        data = Reference(ws, min_col=3, min_row=6, max_row=row-1)
        labels = Reference(ws, min_col=1, min_row=7, max_row=row-1)
        
        chart2.add_data(data, titles_from_data=False)
        chart2.set_categories(labels)
        
        # Add data labels
        chart2.dataLabels = DataLabelList()
        chart2.dataLabels.showPercent = True
        chart2.dataLabels.showLeaderLines = True
        
        ws.add_chart(chart2, f"F{chart_row}")
        
        # Add KPI indicators
        kpi_row = chart_row + 15
        
        kpis = [
            ("Total Dashboards", len([d for d in all_data.values() if d])),
            ("Total Records", sum(len(d) for d in all_data.values())),
            ("Total Engagement", total_engagement),
            ("Avg Confidence", self._calculate_avg_confidence(all_data))
        ]
        
        for i, (label, value) in enumerate(kpis):
            col = i * 2 + 1
            ws.cell(row=kpi_row, column=col, value=label).font = self.bold_font
            ws.cell(row=kpi_row + 1, column=col, value=value).font = Font(size=14, bold=True, color=self.primary_color)
            ws.merge_cells(start_row=kpi_row + 1, start_column=col, end_row=kpi_row + 1, end_column=col + 1)
    
    def _format_dashboard_sheet(self, worksheet, df: pd.DataFrame, dashboard_type: str):
        """Format individual dashboard sheet"""
        # Apply header formatting
        max_column = worksheet.max_column
        
        for cell in worksheet[1]:
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.center_alignment
            cell.border = self.thin_border
        
        # Format data rows
        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, 
                                      min_col=1, max_col=max_column):
            for cell in row:
                cell.font = self.data_font
                cell.border = self.thin_border
                
                # Alternate row coloring
                if cell.row % 2 == 0:
                    cell.fill = self.alt_row_fill
                
                # Right-align numeric columns
                if isinstance(cell.value, (int, float)):
                    cell.alignment = self.right_alignment
                    cell.number_format = '#,##0'  # Format numbers with commas
                else:
                    cell.alignment = self.left_alignment
        
        # Add sheet title
        worksheet.insert_rows(1, 3)
        
        # Main title
        title_row = 1
        worksheet.merge_cells(f'A{title_row}:{get_column_letter(max_column)}{title_row}')
        title_cell = worksheet.cell(row=title_row, column=1)
        title_cell.value = f"{self._format_dashboard_name(dashboard_type)} ANALYSIS"
        title_cell.font = Font(bold=True, size=16, color=self.primary_color)
        title_cell.alignment = self.center_alignment
        
        # Subtitle
        subtitle_row = 2
        worksheet.merge_cells(f'A{subtitle_row}:{get_column_letter(max_column)}{subtitle_row}')
        subtitle_cell = worksheet.cell(row=subtitle_row, column=1)
        subtitle_cell.value = "Detailed Performance Metrics and Analytics"
        subtitle_cell.font = Font(size=12, color=self.secondary_color)
        subtitle_cell.alignment = self.center_alignment
        
        # Freeze header row
        worksheet.freeze_panes = 'A4'
    
    def _add_dashboard_summary(self, worksheet, df: pd.DataFrame, dashboard_type: str):
        """Add summary statistics to dashboard sheet"""
        if len(df) == 0:
            return
        
        start_row = worksheet.max_row + 3
        
        # Summary section title
        summary_title = f"{self._format_dashboard_name(dashboard_type)} - Performance Summary"
        worksheet.cell(row=start_row, column=1, value=summary_title)
        worksheet.cell(row=start_row, column=1).font = Font(bold=True, size=14, color=self.primary_color)
        worksheet.merge_cells(f'A{start_row}:E{start_row}')
        
        # Summary headers
        headers_row = start_row + 2
        headers = ['Metric', 'Total', 'Average', 'Maximum', 'Minimum', 'Std Dev']
        
        for col, header in enumerate(headers, 1):
            cell = worksheet.cell(row=headers_row, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.center_alignment
            cell.border = self.thin_border
        
        # Identify numeric columns for summary
        numeric_cols = df.select_dtypes(include=[int, float]).columns.tolist()
        
        if numeric_cols:
            data_row = headers_row + 1
            
            for col in numeric_cols[:8]:  # Limit to 8 key metrics
                # Metric name
                worksheet.cell(row=data_row, column=1, 
                              value=col.replace('_', ' ').title()).font = self.bold_font
                
                # Statistics
                worksheet.cell(row=data_row, column=2, value=df[col].sum())
                worksheet.cell(row=data_row, column=3, value=df[col].mean())
                worksheet.cell(row=data_row, column=4, value=df[col].max())
                worksheet.cell(row=data_row, column=5, value=df[col].min())
                worksheet.cell(row=data_row, column=6, value=df[col].std())
                
                # Format numeric cells
                for col_idx in range(2, 7):
                    cell = worksheet.cell(row=data_row, column=col_idx)
                    cell.number_format = '#,##0.00'
                    cell.font = self.number_font
                    cell.border = self.thin_border
                    cell.alignment = self.right_alignment
                
                data_row += 1
            
            # Highlight top performers
            if dashboard_type == 'social_media' and 'engagement' in df.columns:
                self._add_top_performers(worksheet, df, data_row + 2)
    
    def _add_charts_to_sheet(self, worksheet, df: pd.DataFrame, dashboard_type: str):
        """Add charts to dashboard sheet"""
        if len(df) < 2:  # Need at least 2 data points for charts
            return
        
        chart_start_row = worksheet.max_row + 3
        
        if dashboard_type == 'social_media' and 'platform' in df.columns:
            # Platform comparison chart
            platform_data = df.groupby('platform').agg({
                'reach_views': 'sum',
                'engagement': 'sum'
            }).reset_index()
            
            # Write chart data
            chart_data_row = chart_start_row
            worksheet.cell(row=chart_data_row, column=1, value="Platform")
            worksheet.cell(row=chart_data_row, column=2, value="Reach/Views")
            worksheet.cell(row=chart_data_row, column=3, value="Engagement")
            
            for i, row in platform_data.iterrows():
                data_row = chart_data_row + i + 1
                worksheet.cell(row=data_row, column=1, value=row['platform'])
                worksheet.cell(row=data_row, column=2, value=row['reach_views'])
                worksheet.cell(row=data_row, column=3, value=row['engagement'])
            
            # Create chart
            chart = BarChart()
            chart.type = "col"
            chart.style = 10
            chart.title = "Platform Performance Comparison"
            chart.y_axis.title = "Count"
            chart.x_axis.title = "Platform"
            
            data = Reference(worksheet, 
                           min_col=2, 
                           min_row=chart_data_row, 
                           max_col=3, 
                           max_row=chart_data_row + len(platform_data))
            categories = Reference(worksheet, 
                                 min_col=1, 
                                 min_row=chart_data_row + 1, 
                                 max_row=chart_data_row + len(platform_data))
            
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(categories)
            
            # Position chart
            chart_cell = f"F{chart_start_row}"
            worksheet.add_chart(chart, chart_cell)
    
    def _create_data_quality_sheet(self, workbook, all_data: Dict[str, List]):
        """Create data quality assessment sheet"""
        ws = workbook.create_sheet(title="Data_Quality")
        
        # Title
        ws.merge_cells('A1:H1')
        ws['A1'] = "DATA QUALITY ASSESSMENT REPORT"
        ws['A1'].font = Font(bold=True, size=16, color=self.primary_color)
        ws['A1'].alignment = self.center_alignment
        
        # Data quality metrics table
        headers = ["Dashboard", "Records", "Completeness", "Consistency", "Accuracy", 
                  "Overall Score", "Status"]
        
        row = 3
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.center_alignment
        
        row += 1
        
        for dashboard_type, data_list in all_data.items():
            if data_list:
                df = pd.DataFrame(data_list)
                
                # Calculate quality metrics
                completeness = self._calculate_completeness(df)
                consistency = self._calculate_consistency(df)
                accuracy = self._calculate_accuracy(df, dashboard_type)
                
                overall_score = (completeness + consistency + accuracy) / 3
                status = self._get_quality_status(overall_score)
                
                values = [
                    self._format_dashboard_name(dashboard_type),
                    len(df),
                    f"{completeness:.1%}",
                    f"{consistency:.1%}",
                    f"{accuracy:.1%}",
                    f"{overall_score:.1%}",
                    status
                ]
                
                for col, value in enumerate(values, 1):
                    cell = ws.cell(row=row, column=col, value=value)
                    cell.font = self.data_font
                    cell.border = self.thin_border
                    cell.alignment = self.center_alignment
                    
                    # Color code status
                    if col == 7:  # Status column
                        if status == "Excellent":
                            cell.font = Font(color=self.success_color, bold=True)
                        elif status == "Good":
                            cell.font = Font(color=self.success_color)
                        elif status == "Fair":
                            cell.font = Font(color=self.warning_color)
                        else:
                            cell.font = Font(color=self.danger_color)
                
                row += 1
        
        # Add recommendations
        rec_row = row + 2
        ws.merge_cells(f'A{rec_row}:H{rec_row}')
        ws.cell(row=rec_row, column=1, value="RECOMMENDATIONS").font = Font(bold=True, size=14, color=self.primary_color)
        
        recommendations = [
            "1. Ensure all required fields are populated in source data",
            "2. Standardize date formats across all data sources",
            "3. Implement data validation rules at point of entry",
            "4. Regularly audit data for consistency and accuracy",
            "5. Establish data quality KPIs and monitoring"
        ]
        
        for i, recommendation in enumerate(recommendations):
            ws.cell(row=rec_row + 2 + i, column=1, value=recommendation).font = self.data_font
    
    def _create_executive_summary(self, workbook, all_data: Dict[str, List]):
        """Create executive summary with insights"""
        ws = workbook.create_sheet(title="Executive_Summary")
        
        # Title
        ws.merge_cells('A1:H1')
        ws['A1'] = "EXECUTIVE SUMMARY - STRATEGIC INSIGHTS"
        ws['A1'].font = Font(bold=True, size=18, color=self.primary_color)
        ws['A1'].alignment = self.center_alignment
        
        # Key Findings
        ws.merge_cells('A3:H3')
        ws['A3'] = "KEY FINDINGS"
        ws['A3'].font = Font(bold=True, size=14, color=self.secondary_color)
        
        findings = self._generate_executive_findings(all_data)
        
        for i, finding in enumerate(findings):
            ws.cell(row=5 + i, column=1, value=f"• {finding}").font = self.data_font
        
        # Recommendations
        rec_row = 5 + len(findings) + 2
        ws.merge_cells(f'A{rec_row}:H{rec_row}')
        ws.cell(row=rec_row, column=1, value="RECOMMENDATIONS").font = Font(bold=True, size=14, color=self.secondary_color)
        
        recommendations = self._generate_recommendations(all_data)
        
        for i, recommendation in enumerate(recommendations):
            ws.cell(row=rec_row + 2 + i, column=1, value=f"{i+1}. {recommendation}").font = self.data_font
        
        # Next Steps
        steps_row = rec_row + len(recommendations) + 4
        ws.merge_cells(f'A{steps_row}:H{steps_row}')
        ws.cell(row=steps_row, column=1, value="NEXT STEPS").font = Font(bold=True, size=14, color=self.secondary_color)
        
        next_steps = [
            "Review detailed dashboard analyses for actionable insights",
            "Implement recommended optimization strategies",
            "Schedule monthly performance review meetings",
            "Update marketing strategies based on data-driven insights",
            "Continue monitoring key performance indicators"
        ]
        
        for i, step in enumerate(next_steps):
            ws.cell(row=steps_row + 2 + i, column=1, value=f"✓ {step}").font = self.data_font
    
    def _generate_executive_findings(self, all_data: Dict[str, List]) -> List[str]:
        """Generate executive findings from data"""
        findings = []
        
        # Analyze social media data
        if 'social_media' in all_data and all_data['social_media']:
            df = pd.DataFrame(all_data['social_media'])
            
            if 'platform' in df.columns and 'engagement' in df.columns:
                platform_engagement = df.groupby('platform')['engagement'].mean()
                best_platform = platform_engagement.idxmax()
                findings.append(f"{best_platform} shows highest average engagement per post")
        
        # Analyze performance marketing
        if 'performance_marketing' in all_data and all_data['performance_marketing']:
            df = pd.DataFrame(all_data['performance_marketing'])
            
            if 'roi' in df.columns:
                avg_roi = df['roi'].mean()
                findings.append(f"Average ROI across campaigns: {avg_roi:.1f}%")
        
        # Analyze KOL engagement
        if 'kol_engagement' in all_data and all_data['kol_engagement']:
            df = pd.DataFrame(all_data['kol_engagement'])
            
            if 'kol_tier' in df.columns:
                tier_dist = df['kol_tier'].value_counts()
                dominant_tier = tier_dist.index[0]
                findings.append(f"{dominant_tier} tier influencers dominate campaign reach")
        
        return findings
    
    def _generate_recommendations(self, all_data: Dict[str, List]) -> List[str]:
        """Generate recommendations based on data analysis"""
        recommendations = []
        
        # Social media recommendations
        if 'social_media' in all_data and all_data['social_media']:
            df = pd.DataFrame(all_data['social_media'])
            
            if 'platform' in df.columns and 'engagement' in df.columns:
                platform_stats = df.groupby('platform').agg({
                    'engagement': 'mean',
                    'reach_views': 'mean'
                })
                
                best_platform = platform_stats['engagement'].idxmax()
                recommendations.append(f"Focus content strategy on {best_platform} for maximum engagement")
        
        # Performance marketing recommendations
        if 'performance_marketing' in all_data and all_data['performance_marketing']:
            df = pd.DataFrame(all_data['performance_marketing'])
            
            if 'ad_type' in df.columns and 'roi' in df.columns:
                ad_type_roi = df.groupby('ad_type')['roi'].mean()
                best_ad_type = ad_type_roi.idxmax()
                recommendations.append(f"Allocate more budget to {best_ad_type} campaigns for better ROI")
        
        return recommendations
    
    def _add_top_performers(self, worksheet, df: pd.DataFrame, start_row: int):
        """Add top performers table"""
        if 'post_title' in df.columns and 'engagement' in df.columns:
            top_posts = df.nlargest(5, 'engagement')[['post_title', 'engagement', 'reach_views']]
            
            worksheet.cell(row=start_row, column=1, value="Top Performing Posts").font = Font(bold=True, size=12)
            worksheet.merge_cells(f'A{start_row}:C{start_row}')
            
            headers = ['Post Title', 'Engagement', 'Reach']
            for col, header in enumerate(headers, 1):
                cell = worksheet.cell(row=start_row + 2, column=col, value=header)
                cell.font = self.header_font
                cell.fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
                cell.alignment = self.center_alignment
            
            for i, (_, row) in enumerate(top_posts.iterrows()):
                data_row = start_row + 3 + i
                worksheet.cell(row=data_row, column=1, value=row['post_title'][:50])
                worksheet.cell(row=data_row, column=2, value=row['engagement'])
                worksheet.cell(row=data_row, column=3, value=row['reach_views'])
    
    def _calculate_total_engagement(self, df: pd.DataFrame) -> int:
        """Calculate total engagement"""
        engagement_metrics = ['engagement', 'likes', 'shares', 'comments', 'saved', 'views']
        total = 0
        
        for metric in engagement_metrics:
            if metric in df.columns:
                total += df[metric].sum()
        
        return total
    
    def _calculate_avg_confidence(self, all_data: Dict[str, List]) -> float:
        """Calculate average confidence score"""
        total_confidence = 0
        count = 0
        
        for data_list in all_data.values():
            if data_list:
                df = pd.DataFrame(data_list)
                if 'confidence_score' in df.columns:
                    total_confidence += df['confidence_score'].sum()
                    count += len(df)
        
        return total_confidence / count if count > 0 else 0.0
    
    def _calculate_data_quality(self, df: pd.DataFrame, dashboard_type: str) -> float:
        """Calculate data quality score"""
        if len(df) == 0:
            return 0.0
        
        # Completeness
        numeric_cols = df.select_dtypes(include=[int, float]).columns.tolist()
        completeness = 0.0
        if numeric_cols:
            total_cells = len(df) * len(numeric_cols)
            non_null_cells = sum(df[col].notna().sum() for col in numeric_cols)
            completeness = non_null_cells / total_cells
        
        # Consistency (check for outliers)
        consistency = 1.0
        if numeric_cols:
            for col in numeric_cols[:3]:  # Check first 3 numeric columns
                if df[col].std() > df[col].mean() * 2:  # High variance
                    consistency *= 0.8
        
        # Accuracy (based on validation if available)
        accuracy = 0.8  # Default
        
        if 'confidence_score' in df.columns:
            accuracy = df['confidence_score'].mean()
        elif 'validation_status' in df.columns:
            valid_count = (df['validation_status'] == 'valid').sum()
            accuracy = valid_count / len(df)
        
        # Weighted average
        quality_score = (completeness * 0.3 + consistency * 0.3 + accuracy * 0.4)
        
        return quality_score
    
    def _get_quality_status(self, score: float) -> str:
        """Get quality status based on score"""
        if score >= 0.9:
            return "Excellent"
        elif score >= 0.7:
            return "Good"
        elif score >= 0.5:
            return "Fair"
        else:
            return "Poor"
    
    def _calculate_completeness(self, df: pd.DataFrame) -> float:
        """Calculate data completeness"""
        if len(df) == 0:
            return 0.0
        
        numeric_cols = df.select_dtypes(include=[int, float]).columns.tolist()
        if not numeric_cols:
            return 0.0
        
        total_cells = len(df) * len(numeric_cols)
        non_null_cells = sum(df[col].notna().sum() for col in numeric_cols)
        
        return non_null_cells / total_cells
    
    def _calculate_consistency(self, df: pd.DataFrame) -> float:
        """Calculate data consistency"""
        if len(df) < 2:
            return 1.0
        
        numeric_cols = df.select_dtypes(include=[int, float]).columns.tolist()
        if not numeric_cols:
            return 1.0
        
        consistency_score = 1.0
        
        for col in numeric_cols[:3]:  # Check first 3 columns
            if df[col].std() == 0:
                continue  # No variation
            
            # Check coefficient of variation
            cv = df[col].std() / df[col].mean()
            if cv > 1.0:  # High variation
                consistency_score *= 0.9
        
        return consistency_score
    
    def _calculate_accuracy(self, df: pd.DataFrame, dashboard_type: str) -> float:
        """Calculate data accuracy"""
        if len(df) == 0:
            return 0.0
        
        # Check for validation status or confidence scores
        if 'confidence_score' in df.columns:
            return df['confidence_score'].mean()
        elif 'validation_status' in df.columns:
            valid_count = (df['validation_status'] == 'valid').sum()
            return valid_count / len(df)
        
        # Default accuracy based on dashboard type
        default_scores = {
            'social_media': 0.85,
            'performance_marketing': 0.80,
            'kol_engagement': 0.75,
            'community_marketing': 0.70,
            'promotion_posts': 0.80
        }
        
        return default_scores.get(dashboard_type, 0.75)
    
    def _get_date_range_from_df(self, df: pd.DataFrame) -> str:
        """Extract date range from dataframe"""
        date_columns = ['post_date', 'video_date', 'created_at', 'date']
        
        for date_col in date_columns:
            if date_col in df.columns:
                try:
                    dates = pd.to_datetime(df[date_col], errors='coerce')
                    valid_dates = dates.dropna()
                    
                    if not valid_dates.empty:
                        min_date = valid_dates.min().strftime('%Y-%m-%d')
                        max_date = valid_dates.max().strftime('%Y-%m-%d')
                        return f"{min_date} to {max_date}"
                except:
                    continue
        
        return "N/A"
    
    def _format_sheet_name(self, dashboard_type: str) -> str:
        """Format sheet name for Excel (max 31 characters)"""
        names = {
            'social_media': 'Social_Media',
            'community_marketing': 'Community_Marketing',
            'kol_engagement': 'KOL_Engagement',
            'performance_marketing': 'Performance_Marketing',
            'promotion_posts': 'Promotion_Posts'
        }
        return names.get(dashboard_type, dashboard_type[:31])
    
    def _format_dashboard_name(self, dashboard_type: str) -> str:
        """Format dashboard name for display"""
        names = {
            'social_media': 'Social Media Management',
            'community_marketing': 'Community Marketing',
            'kol_engagement': 'KOL Engagement',
            'performance_marketing': 'Performance Marketing',
            'promotion_posts': 'Promotion Posts'
        }
        return names.get(dashboard_type, dashboard_type.replace('_', ' ').title())
    
    def _auto_adjust_columns(self, worksheet):
        """Auto-adjust column widths"""
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width